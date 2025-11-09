import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import time

# =========================
# Configurazione pagina
# =========================
st.set_page_config(page_title="Report Attivit√† Clienti", page_icon="üìä")

st.title("üìä Generatore Report Attivit√† Clienti")
st.caption("Versione 1.1 ‚Äì EdiliziAcrobatica S.p.A. (logica 'Da riassegnare' aggiornata)")

# =========================
# Upload dei file
# =========================
file_att = st.file_uploader("üìò Carica il file delle attivit√† (.xlsx)", type=["xlsx"])
file_tab = st.file_uploader("üìó Carica il file Tabella Clienti (.xlsx)", type=["xlsx"])

# =========================
# Avvio elaborazione
# =========================
if file_att and file_tab:
    start_time = time.time()
    st.info("‚è≥ Elaborazione in corso...")

    # 1. Leggo i file
    att = pd.read_excel(file_att)
    tab_raw = pd.read_excel(file_tab, header=None, skiprows=3)
    tab_raw.columns = tab_raw.iloc[0]
    tab = tab_raw.drop(0).reset_index(drop=True)

    # normalizzo nome colonna Macroarea
    tab = tab.rename(columns={"macroarea": "Macroarea"})

    # 2. Normalizzazione nomi
    def normalize_name(x):
        if pd.isna(x): return ""
        x = str(x).lower().replace(".", " ").replace("*", " ").replace(",", " ")
        return " ".join(x.split())

    att["NomeSoggetto_n"] = att["NomeSoggetto"].apply(normalize_name)
    tab["Cliente_n"] = tab["Cliente"].apply(normalize_name)

    # Tipo (colonna P)
    if "Tipo" in tab.columns:
        def fix_tipo(x):
            x = str(x).strip().capitalize()
            if x.lower().startswith("amministrator"):
                return "Amministratori"
            return x
        tab["Tipo"] = tab["Tipo"].apply(fix_tipo)
    else:
        tab["Tipo"] = "Amministratori"

    # Priorit√†
    priorita = {
        "04 RICHIESTE": 1,
        "06 PREVENTIVI": 2,
        "03 INCONTRI": 3,
        "07 DELIBERE": 4,
        "05 SOPRALLUOGHI": 5,
        "01 TELEFONATE": 6,
        "02 APPUNTAMENTI": 7
    }
    att["Priorita"] = att["Classe Attivit√†"].map(priorita).fillna(999)

    # Data corrente
    oggi = datetime.now()
    anno_corrente, mese_corrente = oggi.year, oggi.month

    # 3. Match principale
    righe_output = []
    for _, r in tab.iterrows():
        cliente_norm = r["Cliente_n"]
        tipo_cli = r["Tipo"]
        sede_cli = r.get("Sede", "")
        resp_gest = r.get("Responsabile", "")

        att_cli = att[att["NomeSoggetto_n"] == cliente_norm]
        if att_cli.empty and cliente_norm:
            invertito = " ".join(cliente_norm.split()[::-1])
            att_cli = att[att["NomeSoggetto_n"] == invertito]

        if not att_cli.empty:
            att_cli = att_cli.sort_values(["Anno", "Mese", "Priorita"]).iloc[-1]
            anno_att, mese_att = att_cli["Anno"], att_cli["Mese"]

            # üßÆ nuova logica calcolo "Da riassegnare"
            try:
                anno_att = int(anno_att)
                mese_att = int(float(mese_att))
                diff_mesi = (anno_corrente - anno_att) * 12 + (mese_corrente - mese_att)
                da_ria = "S√¨" if diff_mesi > 2 else "No"
            except Exception:
                da_ria = "S√¨"

            righe_output.append({
                "Sede": sede_cli,
                "Responsabile gestionale": resp_gest,
                "Cliente": r["Cliente"],
                "Anno": anno_att,
                "Mese": mese_att,
                "Ultima attivit√†": att_cli["Classe Attivit√†"],
                "Da riassegnare": da_ria,
                "PREVENTIVATO‚Ç¨": r.get("PREVENTIVATO‚Ç¨", ""),
                "DELIBERATO‚Ç¨": r.get("DELIBERATO‚Ç¨", ""),
                "FATTURATO‚Ç¨": r.get("FATTURATO‚Ç¨", ""),
                "INCASSATO‚Ç¨": r.get("INCASSATO‚Ç¨", ""),
                "Tipo": tipo_cli
            })
        else:
            righe_output.append({
                "Sede": sede_cli,
                "Responsabile gestionale": resp_gest,
                "Cliente": r["Cliente"],
                "Anno": "",
                "Mese": "",
                "Ultima attivit√†": "",
                "Da riassegnare": "S√¨",
                "PREVENTIVATO‚Ç¨": r.get("PREVENTIVATO‚Ç¨", ""),
                "DELIBERATO‚Ç¨": r.get("DELIBERATO‚Ç¨", ""),
                "FATTURATO‚Ç¨": r.get("FATTURATO‚Ç¨", ""),
                "INCASSATO‚Ç¨": r.get("INCASSATO‚Ç¨", ""),
                "Tipo": tipo_cli
            })

    # 4. Attivit√† senza cliente ‚Üí Amministratori
    clienti_norm = set(tab["Cliente_n"].dropna().unique())
    att_no_match = att[~att["NomeSoggetto_n"].isin(clienti_norm)].copy()

    if not att_no_match.empty:
        att_no_match = (
            att_no_match.sort_values(["Anno", "Mese", "Priorita"])
            .groupby("NomeSoggetto", as_index=False)
            .last()
        )

        def da_ria_att(row):
            try:
                anno = int(row["Anno"])
                mese = int(float(row["Mese"]))
                diff = (anno_corrente - anno) * 12 + (mese_corrente - mese)
                return "S√¨" if diff > 2 else "No"
            except Exception:
                return "S√¨"

        att_no_match["Da riassegnare"] = att_no_match.apply(da_ria_att, axis=1)
        att_no_match["Sede"] = att_no_match["Sede"]
        att_no_match["Responsabile gestionale"] = att_no_match["Responsabile"]
        att_no_match["Cliente"] = att_no_match["NomeSoggetto"]
        att_no_match["Ultima attivit√†"] = att_no_match["Classe Attivit√†"]
        att_no_match["Tipo"] = "Amministratori"

        for c in ["PREVENTIVATO‚Ç¨","DELIBERATO‚Ç¨","FATTURATO‚Ç¨","INCASSATO‚Ç¨"]:
            att_no_match[c] = ""

        righe_output.extend(att_no_match[[
            "Sede","Responsabile gestionale","Cliente","Anno","Mese","Ultima attivit√†",
            "Da riassegnare","PREVENTIVATO‚Ç¨","DELIBERATO‚Ç¨","FATTURATO‚Ç¨","INCASSATO‚Ç¨","Tipo"
        ]].to_dict(orient="records"))

    # 5. DataFrame finale
    database = pd.DataFrame(righe_output).replace({np.nan: ""})

    def to_float_euro(x):
        if pd.isna(x) or str(x).strip() == "":
            return np.nan
        x = str(x).replace("‚Ç¨", "").replace(" ", "")
        if "," in x and "." in x:
            x = x.replace(".", "").replace(",", ".")
        elif "," in x:
            x = x.replace(",", ".")
        try:
            return float(x)
        except ValueError:
            return np.nan

    def format_euro(x):
        if pd.isna(x) or x == "":
            return ""
        try:
            return f"‚Ç¨ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return str(x)

    for c in ["PREVENTIVATO‚Ç¨","DELIBERATO‚Ç¨","FATTURATO‚Ç¨","INCASSATO‚Ç¨"]:
        if c in database.columns:
            database[c] = database[c].apply(to_float_euro).apply(format_euro)

    # 6. Esporta Excel
    output = BytesIO()
    col_order = [
        "Sede","Responsabile gestionale","Cliente","Anno","Mese",
        "Ultima attivit√†","Da riassegnare",
        "PREVENTIVATO‚Ç¨","DELIBERATO‚Ç¨","FATTURATO‚Ç¨","INCASSATO‚Ç¨"
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        database.to_excel(writer, sheet_name="Database", index=False)
        for tipo, grp in sorted(database.groupby("Tipo"), key=lambda x: str(x[0])):
            nome = str(tipo).strip().capitalize() or "Senzatipo"
            grp[col_order].sort_values("Cliente").to_excel(writer, sheet_name=nome, index=False)

    # 7. Formattazione
    output.seek(0)
    wb = load_workbook(output)
    thin = Side(border_style="thin", color="D9D9D9")
    header_fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="A6F3A6", end_color="A6F3A6", fill_type="solid")

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = alt_fill
                if cell.value == "S√¨":
                    cell.fill = red_fill
                elif cell.value == "No":
                    cell.fill = green_fill
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)

    if "Amministratori" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Amministratori")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    elapsed = time.time() - start_time
    minuti = int(elapsed // 60)
    secondi = int(elapsed % 60)

    st.success(f"‚úÖ Report completato in {minuti} min {secondi} sec!")
    st.download_button(
        label="üì• Scarica report_attivita_clienti.xlsx",
        data=buf,
        file_name="report_attivita_clienti.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
